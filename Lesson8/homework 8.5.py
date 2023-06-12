import csv
import openpyxl

with open('data.csv', newline='') as csvfile:
    data = csv.reader(csvfile, delimiter='\t', quotechar='|')

    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A2'] = 100
    i = 1
    for line in data:
        sheet.cell(row=2, column=i).value = line[0]
        sheet.cell(row=3, column=i).value = line[1]
        sheet.cell(row=4, column=i).value = line[3]
        i += 1
    for i in range(1, 6):
        sheet.cell(row=1, column=i+1).value = f'Person {i}'
    book.save('data.xlsx')
    book.close()

